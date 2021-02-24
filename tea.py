# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file '20200421.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!

from hun1_png import img as tea_1
from tea_logo_jpg import img as tea
from hun_logo1_jpg import img as hun
import os
os.environ['BLINKA_FT232H'] = '1'
from PyQt5.QtGui import QPalette, QColor, QImage, QPixmap, QPainter, QPen, QGuiApplication
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
import cv2
import numpy as np
import time
import datetime
##import board
import time
##import digitalio
from PyQt5.QtWidgets import QFileDialog, QWidget, QApplication, QLabel, QDialog, QProgressBar, QPushButton, QMessageBox
from PyQt5.QtCore import QRect, Qt
import openpyxl  
import base64
from openpyxl.styles import  PatternFill

TIME_LIMIT = 100

class Actions(QDialog):
    """
    Simple dialog that consists of a Progress Bar and a Button.
    Clicking on the button results in the start of a timer and
    updates the progress bar.
    """
    def __init__(self):
        super().__init__()
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('開啟中')
        self.resize(463, 152)
        self.verticalLayoutWidget = QtWidgets.QWidget(self)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(0, 0, 461, 151))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.verticalLayoutWidget)

        image_tea1 = base64.b64decode(tea_1)
        pix = QPixmap()
        pix.loadFromData(image_tea1)
        self.label.setGeometry(1,1,114,83)
        self.label.setScaledContents (False)
        self.label.setPixmap(pix)

        self.label.setText("")
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.progressBar = QtWidgets.QProgressBar(self.verticalLayoutWidget)
        self.progressBar.setMaximum(100)
        self.progressBar.setProperty("value", 24)
        self.progressBar.setObjectName("progressBar")
        self.verticalLayout.addWidget(self.progressBar)
        self.show()
        count = 0
        while count < TIME_LIMIT:
            count += 1
            time.sleep(0.01)
            self.progressBar.setValue(count)

    def onButtonClick(self):
        count = 0
        while count < TIME_LIMIT:
            count += 1
            time.sleep(0.01)
            self.progress.setValue(count)


class Ui_MainWindow(object):
    def __init__(self, MainWindow):

        app.aboutToQuit.connect(self.closeEvent)
        
        self.right = 0
        self.timer_camera = QtCore.QTimer()
        
        self.setupUi(MainWindow)
        self.retranslateUi(MainWindow)
                
        self.cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
       # cv2.imwrite(r'C:\Users\a0913\Desktop\82.jpg',self.image)
        self.CAM_NUM = 0
        self.CAM_SET_COUNT = 0
        self.mouse_is_pressing = False
        self.start_x, self.start_y = -1, -1
        self.button_count = 1
        self.path_success_count = 0
        self.slot_init()
     ##   led = digitalio.DigitalInOut(board.C0)
     ##   led.direction = digitalio.Direction.OUTPUT
     ##   led.value = True

        print ("__start__")

    def setupUi(self, mainWindow):
        mainWindow.setObjectName("mainWindow")
        mainWindow.resize(1283, 841)
        font = QtGui.QFont()
        font.setFamily("Agency FB")
        font.setBold(True)
        font.setWeight(75)
        mainWindow.setFont(font)
        mainWindow.setStyleSheet("")
        self.centralwidget = QtWidgets.QWidget(mainWindow)
        self.centralwidget.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.centralwidget.sizePolicy().hasHeightForWidth())
        self.centralwidget.setSizePolicy(sizePolicy)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.gridLayout_5 = QtWidgets.QGridLayout()
        self.gridLayout_5.setSizeConstraint(QtWidgets.QLayout.SetFixedSize)
        self.gridLayout_5.setContentsMargins(0, 0, 0, 5)
        self.gridLayout_5.setHorizontalSpacing(6)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.label_27 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_27.setFont(font)
        self.label_27.setStyleSheet("background-color: rgb(0, 170, 0);\n"
"color: rgb(255, 255, 255);")
        self.label_27.setAlignment(QtCore.Qt.AlignCenter)
        self.label_27.setObjectName("label_27")
        self.gridLayout_5.addWidget(self.label_27, 0, 1, 1, 1)
        self.label_26 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_26.setFont(font)
        self.label_26.setStyleSheet("background-color: rgb(10, 255, 26);")
        self.label_26.setAlignment(QtCore.Qt.AlignCenter)
        self.label_26.setObjectName("label_26")
        self.gridLayout_5.addWidget(self.label_26, 1, 1, 1, 1)
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setText("")
        self.label.setObjectName("label")

        image_tea = base64.b64decode(tea)
        pix = QPixmap()
        pix.loadFromData(image_tea)
        self.label.setGeometry(1,1,114,83)
       # self.label.setStyleSheet("border: 2px solid red")
        self.label.setScaledContents (False)
        self.label.setPixmap(pix)

        self.gridLayout_5.addWidget(self.label, 0, 0, 2, 1)
        self.gridLayout_5.setColumnStretch(0, 1)
        self.gridLayout_5.setColumnStretch(1, 10)
        self.verticalLayout.addLayout(self.gridLayout_5)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 127, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(63, 63, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Midlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Dark, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 170))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Mid, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.BrightText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Shadow, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 127, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 220))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ToolTipBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ToolTipText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 127, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(63, 63, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Midlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Dark, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 170))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Mid, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.BrightText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Shadow, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 127, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 220))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ToolTipBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ToolTipText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(127, 127, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Light, brush)
        brush = QtGui.QBrush(QtGui.QColor(63, 63, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Midlight, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Dark, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 170))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Mid, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.BrightText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 127))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(100, 255, 139))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Shadow, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.AlternateBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 220))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ToolTipBase, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ToolTipText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0, 128))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        self.pushButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("border:2px groove gray;border-radius:10px;padding:15px 4px;\n"
"background-color: rgb(100, 255, 139);")
        icon = QtGui.QIcon.fromTheme("123")
        self.pushButton.setIcon(icon)
        self.pushButton.setAutoDefault(False)
        self.pushButton.setDefault(False)
        self.pushButton.setFlat(False)
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("border:2px groove gray;border-radius:10px;padding:15px 4px;\n"
"background-color: rgb(100, 255, 139);")
        self.pushButton_2.setObjectName("pushButton_2")
        self.horizontalLayout.addWidget(self.pushButton_2)
        self.pushButton_6 = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setStyleSheet("border:2px groove gray;border-radius:10px;padding:15px 4px;\n"
"background-color: rgb(186, 255, 139);")
        self.pushButton_6.setObjectName("pushButton_6")
        self.horizontalLayout.addWidget(self.pushButton_6)
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setStyleSheet("border:2px groove gray;border-radius:10px;padding:15px 4px;\n"
"background-color: rgb(186, 255, 139);")
        self.pushButton_4.setObjectName("pushButton_4")
        self.horizontalLayout.addWidget(self.pushButton_4)
        self.pushButton_5 = QtWidgets.QPushButton(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setStyleSheet("border:2px groove gray;border-radius:10px;padding:15px 4px;\n"
"background-color: rgb(186, 255, 139);")
        self.pushButton_5.setObjectName("pushButton_5")
        self.horizontalLayout.addWidget(self.pushButton_5)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setHorizontalSpacing(7)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_10 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_10.setScaledContents(False)
        self.label_10.setAlignment(QtCore.Qt.AlignCenter)
        self.label_10.setObjectName("label_10")
        self.gridLayout_2.addWidget(self.label_10, 21, 0, 1, 1)
        self.lcdNumber_6 = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber_6.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber_6.setLineWidth(1)
        self.lcdNumber_6.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber_6.setObjectName("lcdNumber_6")
        self.gridLayout_2.addWidget(self.lcdNumber_6, 23, 1, 1, 4)
        self.label_19 = QtWidgets.QLabel(self.centralwidget)

        palette = QPalette()
        palette.setColor(QPalette.Window,QColor.fromHsv(116,0,116))
        #palette.setColor(QPalette.WindowText,QColor(255,0,0))
        self.label_19.setPalette(palette)

        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_19.setFont(font)
        self.label_19.setAutoFillBackground(True)
        self.label_19.setStyleSheet("")
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setObjectName("label_19")
        self.gridLayout_2.addWidget(self.label_19, 19, 1, 1, 4)
        self.lcdNumber = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber.setAutoFillBackground(False)
        self.lcdNumber.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber.setFrameShape(QtWidgets.QFrame.Box)
        self.lcdNumber.setFrameShadow(QtWidgets.QFrame.Raised)
        self.lcdNumber.setLineWidth(1)
        self.lcdNumber.setMidLineWidth(0)
        self.lcdNumber.setSmallDecimalPoint(False)
        self.lcdNumber.setDigitCount(5)
        self.lcdNumber.setMode(QtWidgets.QLCDNumber.Dec)
        self.lcdNumber.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber.setObjectName("lcdNumber")
        self.gridLayout_2.addWidget(self.lcdNumber, 12, 1, 1, 4)
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_5.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setObjectName("label_5")
        self.gridLayout_2.addWidget(self.label_5, 1, 0, 1, 1)
        self.horizontalSlider = QtWidgets.QSlider(self.centralwidget)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 249, 242))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.horizontalSlider.setPalette(palette)
        self.horizontalSlider.setMinimum(0)
        self.horizontalSlider.setMaximum(255)
        self.horizontalSlider.setSliderPosition(255)
        self.horizontalSlider.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider.setTickPosition(QtWidgets.QSlider.TicksBelow)
        self.horizontalSlider.setTickInterval(0)
        self.horizontalSlider.setObjectName("horizontalSlider")
        self.gridLayout_2.addWidget(self.horizontalSlider, 18, 1, 1, 4)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.horizontalSlider_4 = QtWidgets.QSlider(self.centralwidget)
        self.horizontalSlider_4.setEnabled(True)
        self.horizontalSlider_4.setStyleSheet("")
        self.horizontalSlider_4.setMinimum(0)
        self.horizontalSlider_4.setMaximum(255)
        self.horizontalSlider_4.setProperty("value", 0)
        self.horizontalSlider_4.setSliderPosition(0)
        self.horizontalSlider_4.setOrientation(QtCore.Qt.Horizontal)
        self.horizontalSlider_4.setInvertedAppearance(False)
        self.horizontalSlider_4.setInvertedControls(False)
        self.horizontalSlider_4.setTickPosition(QtWidgets.QSlider.TicksBelow)
        self.horizontalSlider_4.setObjectName("horizontalSlider_4")
        self.horizontalLayout_3.addWidget(self.horizontalSlider_4)
        self.gridLayout_2.addLayout(self.horizontalLayout_3, 16, 1, 1, 4)
        self.label_21 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_21.setFont(font)
        self.label_21.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_21.setAlignment(QtCore.Qt.AlignCenter)
        self.label_21.setObjectName("label_21")
        self.gridLayout_2.addWidget(self.label_21, 23, 0, 1, 1)
        self.lineEdit_5 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_5.setFrame(True)
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.gridLayout_2.addWidget(self.lineEdit_5, 0, 4, 1, 1)
        self.label_18 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_18.setFont(font)
        self.label_18.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setObjectName("label_18")
        self.gridLayout_2.addWidget(self.label_18, 19, 0, 1, 1)
        self.lineEdit_7 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.gridLayout_2.addWidget(self.lineEdit_7, 1, 4, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setAutoFillBackground(False)
        self.label_8.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.gridLayout_2.addWidget(self.label_8, 17, 0, 1, 1)
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("Times New Roman")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_4.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setObjectName("label_4")
        self.gridLayout_2.addWidget(self.label_4, 0, 1, 1, 2)
        self.label_9 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.gridLayout_2.addWidget(self.label_9, 4, 0, 1, 1)
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout_2.addWidget(self.lineEdit, 1, 1, 1, 2)
        self.lcdNumber_4 = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber_4.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber_4.setLineWidth(1)
        self.lcdNumber_4.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber_4.setObjectName("lcdNumber_4")
        self.gridLayout_2.addWidget(self.lcdNumber_4, 21, 1, 1, 4)
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setObjectName("label_7")
        self.gridLayout_2.addWidget(self.label_7, 3, 0, 1, 1)
        self.label_24 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_24.setFont(font)
        self.label_24.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_24.setAlignment(QtCore.Qt.AlignCenter)
        self.label_24.setObjectName("label_24")
        self.gridLayout_2.addWidget(self.label_24, 5, 3, 1, 1)
        self.label_17 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_17.setFont(font)
        self.label_17.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_17.setAlignment(QtCore.Qt.AlignCenter)
        self.label_17.setObjectName("label_17")
        self.gridLayout_2.addWidget(self.label_17, 18, 0, 1, 1)
        self.label_13 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_13.setFont(font)
        self.label_13.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.gridLayout_2.addWidget(self.label_13, 12, 0, 1, 1)
        self.lcdNumber_5 = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber_5.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber_5.setLineWidth(1)
        self.lcdNumber_5.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber_5.setObjectName("lcdNumber_5")
        self.gridLayout_2.addWidget(self.lcdNumber_5, 22, 1, 1, 4)
        self.label_14 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_14.setFont(font)
        self.label_14.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.label_14.setObjectName("label_14")
        self.gridLayout_2.addWidget(self.label_14, 13, 0, 1, 1)
        self.label_22 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_22.setFont(font)
        self.label_22.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_22.setAlignment(QtCore.Qt.AlignCenter)
        self.label_22.setObjectName("label_22")
        self.gridLayout_2.addWidget(self.label_22, 4, 3, 1, 1)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_2.sizePolicy().hasHeightForWidth())
        self.lineEdit_2.setSizePolicy(sizePolicy)
        self.lineEdit_2.setMinimumSize(QtCore.QSize(154, 0))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout_2.addWidget(self.lineEdit_2, 3, 1, 1, 2)
        self.label_20 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_20.setFont(font)
        self.label_20.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.label_20.setObjectName("label_20")
        self.gridLayout_2.addWidget(self.label_20, 22, 0, 1, 1)
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_6.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_6.setLineWidth(1)
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.gridLayout_2.addWidget(self.label_6, 16, 0, 1, 1)
        self.label_12 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_12.setFont(font)
        self.label_12.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.gridLayout_2.addWidget(self.label_12, 0, 3, 1, 1)
        self.label_28 = QtWidgets.QLabel(self.centralwidget)
        self.label_28.setText("")
        self.label_28.setObjectName("label_28")

        image_hun = base64.b64decode(hun)
        pix = QPixmap()
        pix.loadFromData(image_hun)

        self.label_28.setGeometry(1,1,114,83)
        #self.label.setStyleSheet("border: 2px solid red")
        self.label.setScaledContents (False)
        self.label_28.setPixmap(pix)

        self.gridLayout_2.addWidget(self.label_28, 24, 0, 1, 1)
        self.lcdNumber_2 = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber_2.setAutoFillBackground(False)
        self.lcdNumber_2.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber_2.setLineWidth(1)
        self.lcdNumber_2.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber_2.setObjectName("lcdNumber_2")
        self.gridLayout_2.addWidget(self.lcdNumber_2, 13, 1, 1, 4)
        self.lineEdit_4 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.gridLayout_2.addWidget(self.lineEdit_4, 5, 1, 1, 2)
        self.label_16 = QtWidgets.QLabel(self.centralwidget)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        gradient = QtGui.QLinearGradient(0.0, 0.0, 1.0, 0.0)
        gradient.setSpread(QtGui.QGradient.PadSpread)
        gradient.setCoordinateMode(QtGui.QGradient.ObjectBoundingMode)
        gradient.setColorAt(0.0, QtGui.QColor(116, 116, 166))
        gradient.setColorAt(1.0, QtGui.QColor(7, 116, 0))
        brush = QtGui.QBrush(gradient)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        self.label_16.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_16.setFont(font)
        self.label_16.setFocusPolicy(QtCore.Qt.NoFocus)
        self.label_16.setAutoFillBackground(False)
        self.label_16.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(116, 116, 166, 255), stop:1 rgba(7, 116, 0, 255));\n"
"color: rgb(255, 255, 255);")
        self.label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.label_16.setObjectName("label_16")
        self.gridLayout_2.addWidget(self.label_16, 17, 1, 1, 4)
        self.lcdNumber_3 = QtWidgets.QLCDNumber(self.centralwidget)
        self.lcdNumber_3.setAutoFillBackground(False)
        self.lcdNumber_3.setStyleSheet("background-color: rgb(255, 251, 202);\n"
"color: rgb(3, 33, 202);")
        self.lcdNumber_3.setLineWidth(1)
        self.lcdNumber_3.setSegmentStyle(QtWidgets.QLCDNumber.Flat)
        self.lcdNumber_3.setObjectName("lcdNumber_3")
        self.gridLayout_2.addWidget(self.lcdNumber_3, 14, 1, 1, 4)
        self.lineEdit_8 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_8.setObjectName("lineEdit_8")
        self.gridLayout_2.addWidget(self.lineEdit_8, 5, 4, 1, 1)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.gridLayout_2.addWidget(self.lineEdit_3, 4, 1, 1, 2)
        self.lineEdit_9 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.gridLayout_2.addWidget(self.lineEdit_9, 3, 4, 1, 1)
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setItalic(False)
        font.setWeight(75)
        font.setStrikeOut(False)
        font.setKerning(True)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_3.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.label_3.setFrameShadow(QtWidgets.QFrame.Plain)
        self.label_3.setLineWidth(1)
        self.label_3.setMidLineWidth(0)
        self.label_3.setTextFormat(QtCore.Qt.AutoText)
        self.label_3.setScaledContents(False)
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setObjectName("label_3")
        self.gridLayout_2.addWidget(self.label_3, 0, 0, 1, 1)
        self.label_29 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_29.setFont(font)
        self.label_29.setStyleSheet("background-color: rgb(0, 0, 193);\n"
"color: rgb(255, 255, 255);")
        self.label_29.setAlignment(QtCore.Qt.AlignCenter)
        self.label_29.setObjectName("label_29")
        self.gridLayout_2.addWidget(self.label_29, 24, 1, 1, 4)
        self.label_23 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_23.setFont(font)
        self.label_23.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_23.setAlignment(QtCore.Qt.AlignCenter)
        self.label_23.setObjectName("label_23")
        self.gridLayout_2.addWidget(self.label_23, 1, 3, 1, 1)
        self.label_11 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_11.setFont(font)
        self.label_11.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.gridLayout_2.addWidget(self.label_11, 5, 0, 1, 1)
        self.label_25 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_25.setFont(font)
        self.label_25.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_25.setAlignment(QtCore.Qt.AlignCenter)
        self.label_25.setObjectName("label_25")
        self.gridLayout_2.addWidget(self.label_25, 3, 3, 1, 1)
        self.lineEdit_6 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.gridLayout_2.addWidget(self.lineEdit_6, 4, 4, 1, 1)
        self.label_15 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_15.setFont(font)
        self.label_15.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.label_15.setObjectName("label_15")
        self.gridLayout_2.addWidget(self.label_15, 14, 0, 1, 1)
        self.label_30 = QtWidgets.QLabel(self.centralwidget)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_30.setFont(font)
        self.label_30.setStyleSheet("background-color: rgb(255, 251, 202);")
        self.label_30.setAlignment(QtCore.Qt.AlignCenter)
        self.label_30.setObjectName("label_30")
        self.gridLayout_2.addWidget(self.label_30, 20, 0, 1, 1)
        self.label_31 = QtWidgets.QLabel(self.centralwidget)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 240, 223))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 240, 223))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 240, 223))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(32, 240, 223))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.label_31.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_31.setFont(font)
        self.label_31.setAutoFillBackground(True)
        self.label_31.setStyleSheet("background-color: rgb(255, 0, 0);")
        self.label_31.setText("")
        self.label_31.setAlignment(QtCore.Qt.AlignCenter)
        self.label_31.setObjectName("label_31")
        self.gridLayout_2.addWidget(self.label_31, 20, 1, 1, 4)
        self.horizontalLayout_2.addLayout(self.gridLayout_2)
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(1)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_2.sizePolicy().hasHeightForWidth())
        self.label_2.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("微軟正黑體")
        font.setPointSize(22)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setAutoFillBackground(False)
        self.label_2.setStyleSheet("background-color: rgb(0, 0, 0);\n"
"color: rgb(255, 255, 255);")
        self.label_2.setText("")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)
        self.gridLayout.addLayout(self.horizontalLayout_2, 3, 0, 1, 1)
        self.verticalLayout.addLayout(self.gridLayout)
        self.gridLayout_3.addLayout(self.verticalLayout, 2, 0, 1, 1)
        mainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(mainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1283, 25))
        self.menubar.setObjectName("menubar")
        mainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(mainWindow)
        self.statusbar.setObjectName("statusbar")
        mainWindow.setStatusBar(self.statusbar)


        self.retranslateUi(mainWindow)
        QtCore.QMetaObject.connectSlotsByName(mainWindow)

    def retranslateUi(self, mainWindow):
        _translate = QtCore.QCoreApplication.translate
        mainWindow.setWindowTitle(_translate("mainWindow", "MainWindow"))
        self.label_27.setText(_translate("mainWindow", "行政院農業委員會茶業改良場"))
        self.label_26.setText(_translate("mainWindow", "茶葉萎凋系統"))
        self.pushButton.setText(_translate("mainWindow", "讀取先前數據"))
        self.pushButton_2.setText(_translate("mainWindow", "檢測並另存資料及影像"))
        self.pushButton_6.setText(_translate("mainWindow", "檢測並儲存資料及影像"))
        self.pushButton_4.setText(_translate("mainWindow", "關閉光源"))
        self.pushButton_5.setText(_translate("mainWindow", "框選圖片"))
        self.label_10.setText(_translate("mainWindow", "R(紅)"))
        self.label_19.setText(_translate("mainWindow", "0"))
        self.label_5.setText(_translate("mainWindow", "產地"))
        self.label_21.setText(_translate("mainWindow", "B(藍)"))
        self.label_18.setText(_translate("mainWindow", "測量顏色"))
        self.label_8.setText(_translate("mainWindow", "顏色範圍"))
        self.label_4.setText(_translate("mainWindow", "2020-01-02-02-02-02"))
        self.label_9.setText(_translate("mainWindow", "栽培者"))
        self.label_7.setText(_translate("mainWindow", "品種"))
        self.label_24.setText(_translate("mainWindow", "備註2"))
        self.label_17.setText(_translate("mainWindow", "上限(S值)"))
        self.label_13.setText(_translate("mainWindow", "色相(H)"))
        self.label_14.setText(_translate("mainWindow", "飽和度(S)"))
        self.label_22.setText(_translate("mainWindow", "備註1"))
        self.label_20.setText(_translate("mainWindow", "G(綠)"))
        self.label_6.setText(_translate("mainWindow", "下限(S值)"))
        self.label_12.setText(_translate("mainWindow", "重量"))
        self.label_16.setText(_translate("mainWindow", "下限:0 上限:255"))
        self.label_3.setText(_translate("mainWindow", "時間"))
        self.label_29.setText(_translate("mainWindow", "鍠麟機械有限公司"))
        self.label_23.setText(_translate("mainWindow", "目測萎凋程度(%)"))
        self.label_11.setText(_translate("mainWindow", "茶菁品質"))
        self.label_25.setText(_translate("mainWindow", "實際水分含量(%)"))
        self.label_15.setText(_translate("mainWindow", "亮度(V)"))
        self.label_30.setText(_translate("mainWindow", "測量結果"))
        datetime=time.strftime("%Y-%m-%d %H:%M:%S")
        self.label_4.setText(_translate("MainWindow", datetime))

    def slot_init(self):
        self.button_open_camera_click()
        self.timer_camera.timeout.connect(self.show_camera)
        self.pushButton_2.clicked.connect(self.savePhoto)
        self.pushButton.clicked.connect(self.load_data)
        self.pushButton_4.clicked.connect(self.openlight)
        self.pushButton_5.clicked.connect(self.select)
        self.pushButton_6.clicked.connect(self.save)
        self.horizontalSlider.valueChanged.connect(self.label_color_change)
        self.horizontalSlider_4.valueChanged.connect(self.label_color_change)
       # self.horizontalSlider_3.valueChanged.connect(self.label_color_change)

    def load_data(self):
        now_time = time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
        now_time_no_hr = time.strftime('%Y-%m-%d',time.localtime(time.time()))
        now_time_hr = time.strftime('%H-%M-%S',time.localtime(time.time()))
        
        self.fileName,_ = QFileDialog.getOpenFileName(None,'load',now_time_no_hr,'Excel Files(*.xlsx)')
        #print (fileName_load)
        if self.fileName :
            self.path=r'%s'%self.fileName
            wb = openpyxl.load_workbook(self.path)
            ws=wb.active
            nrows_load = ws.max_row
           # print('表格總列數',nrows_load)
            strvalue2 = str(ws.cell(row=nrows_load,column=2).value)
            self.lineEdit.setText(strvalue2)
            strvalue3 = str(ws.cell(row=nrows_load,column=3).value)
            self.lineEdit_2.setText(strvalue3)
            strvalue4 = str(ws.cell(row=nrows_load,column=4).value)
            self.lineEdit_3.setText(strvalue4)
            strvalue5 = str(ws.cell(row=nrows_load,column=5).value)
            self.lineEdit_4.setText(strvalue5)
            strvalue6 = str(ws.cell(row=nrows_load,column=6).value)
            self.lineEdit_5.setText(strvalue6)
            strvalue7 = str(ws.cell(row=nrows_load,column=7).value)
            self.lineEdit_7.setText(strvalue7)
            strvalue8 = str(ws.cell(row=nrows_load,column=8).value)
            self.lineEdit_9.setText(strvalue8)
            strvalue9 = str(ws.cell(row=nrows_load,column=9).value)
            self.lineEdit_6.setText(strvalue9)
            strvalue10 = str(ws.cell(row=nrows_load,column=10).value)
            self.lineEdit_8.setText(strvalue10)
            strvalue16 = str(ws.cell(row=nrows_load,column=16).value)
            self.label_19.setText(strvalue16)
            strvalue21 = str(ws.cell(row=nrows_load,column=21).value)
            self.label_31.setText(strvalue21)

            self.horizontalSlider_4.setSliderPosition(int(ws.cell(row=nrows_load,column=14).value))
            self.horizontalSlider.setSliderPosition(int(ws.cell(row=nrows_load,column=15).value))
            self.path_success = self.path
            self.path_success_count = 1

    def openlight(self):
        led = digitalio.DigitalInOut(board.C0)
        led.direction = digitalio.Direction.OUTPUT
        (self.button_count)=(self.button_count)+1
        if ((self.button_count)%2) == 1:
            led.value = True
        else:
            led.value = False

    def label_color_change(self):
        ##Palette = QPalette()
        #color=QColor.fromHsv(h,s,v,a)
        ##Palette.setColor(QPalette.Window,QColor.fromHsv(116,self.horizontalSlider.value(),116))#设置字体颜色
        ##self.label_16.setPalette(Palette)

        ##Palette = QPalette()
        #color=QColor.fromHsv(h,s,v,a)
        ##Palette.setColor(QPalette.Window,QColor.fromHsv(116,self.horizontalSlider_2.value(),116))#设置字体颜色
        ##self.label_10.setPalette(Palette)
        self.label_16.setText("下限:"+"%d"%self.horizontalSlider_4.value()+" "+"上限:"+"%d"%self.horizontalSlider.value())
        #self.label_19.setText("%d"%(self.horizontalSlider_3.value()))

        #Palette = QPalette()
        #color=QColor.fromHsv(h,s,v,a)
        #Palette.setColor(QPalette.Window,QColor.fromHsv(116,self.S_avg,116))#设置字体颜色
        #self.label_19.setPalette(Palette)

    def button_open_camera_click(self):        
        if self.timer_camera.isActive() == False:
            flag = self.cap.open(self.CAM_NUM)
            if flag == False:
                msg = QtWidgets.QMessageBox.warning(
                self, u"Warning", u"Do you conncet the camera with Raspi?",
                buttons=QtWidgets.QMessageBox.Ok,
                defaultButton=QtWidgets.QMessageBox.Ok)
            else:
                self.timer_camera.start(30)

    def show_camera(self):

        if self.CAM_SET_COUNT == 0:
            self.cap.set(3,2048)
            self.cap.set(4,1536)
            #self.cap.set(3,800)
            #self.cap.set(4,600)
            #self.cap.set(10,100) #BRIGHTNESS ok
           # self.cap.set(11,50) #CONTRAST ok
           # self.cap.set(12,100) #SATURATION ok
            #self.cap.set(13,0) #HUE圖像的色相 ok
            #self.cap.set(15,-1) #EXPOSURE
            self.cap.set(16,1) #指示是否應將圖像轉換為RGB
            self.cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc('M', 'J', 'P', 'G')) 
            self.CAM_SET_COUNT = self.CAM_SET_COUNT + 1
        
        flag, self.image = self.cap.read()
        #print (flag)
        self.image = cv2.resize(self.image,(640, 480)) 
        #self.image=cv2.flip(self.image,1) 
        show = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)
        #print (np.shape(show))
        HSV = cv2.cvtColor(self.image, cv2.COLOR_BGR2HSV)
        
        (R, G, B) = cv2.split(show)
        (H, S, V) = cv2.split(HSV)
        
        self.H_avg=np.mean(H)
        self.S_avg=np.mean(S)
        self.V_avg=np.mean(V)

        self.R_avg=np.mean(R)
        self.G_avg=np.mean(G)
        self.B_avg=np.mean(B)

        _translate = QtCore.QCoreApplication.translate
        showImage = QtGui.QImage(show.data, show.shape[1], show.shape[0], QtGui.QImage.Format_RGB888)
        self.label_2.setPixmap(QtGui.QPixmap.fromImage(showImage))
        self.label_2.setScaledContents(True)
        self.label_2.setCursor(Qt.CrossCursor)
        
        self.lcdNumber.display("%d"%self.H_avg)
        self.lcdNumber_2.display("%d"%self.S_avg)
        self.lcdNumber_3.display("%d"%self.V_avg)

        self.lcdNumber_4.display("%d"%self.R_avg)
        self.lcdNumber_5.display("%d"%self.G_avg)
        self.lcdNumber_6.display("%d"%self.B_avg)

    def savePhoto(self):
       # self.horizontalSlider_3.setSliderPosition(int(self.S_avg))
        now_H = self.H_avg
        now_H = int(now_H)
        now_S = self.S_avg
        now_S = int(now_S)
        now_V = self.V_avg
        now_V = int(now_V)
        now_R = self.R_avg
        now_R = int(now_R)
        now_G = self.G_avg
        now_G = int(now_G)
        now_B = self.B_avg
        now_B = int(now_B)
        now_image=self.image

        self.label_19.setText(str(now_S))
        Palette = QPalette()
        #color=QColor.fromHsv(h,s,v,a)
        Palette.setColor(QPalette.Window,QColor.fromHsv(116,now_S,116))#设置字体颜色
        self.label_19.setPalette(Palette)

        if self.horizontalSlider.value() >  self.horizontalSlider_4.value():
                if self.horizontalSlider_4.value() == 0 :
                        if now_S > self.horizontalSlider.value() :
                                error = (now_S - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)
                        else :
                                error_text = ('萎凋率0%')
                                self.label_31.setText(error_text)
                else :
                        if now_S > self.horizontalSlider.value() :
                                error = (now_S - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)

                        elif now_S >= self.horizontalSlider_4.value() and now_S <= self.horizontalSlider.value() :
                                error_text = ('萎凋率0%')
                                self.label_31.setText(error_text)
                        else :
                                error = (now_S - self.horizontalSlider_4.value())/self.horizontalSlider_4.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)

                if self.timer_camera.isActive() != False:
                        now_time = time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
                        now_time_no_hr = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                        now_time_hr = time.strftime('%H-%M-%S',time.localtime(time.time()))
                        #print(now_time)

                        self.fileName,_ = QFileDialog.getSaveFileName(None,'save',now_time_no_hr,'Excel Files(*.xlsx)')
                        # print (fileName)
                        if self.fileName :
                                if os.path.isfile(self.fileName):
                                        self.__file__=self.fileName
                                        self.pic_dir=os.path.dirname(self.__file__)+'/'+now_time+'.jpg'
                                        #print (pic_dir)
                                        cv2.imencode('.jpg', now_image)[1].tofile(self.pic_dir)
                                        # cv2.putText(self.image, 'The picture and data have saved !',
                                        #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                        #                 cv2.FONT_HERSHEY_SIMPLEX,
                                        #                 1.0, (255, 0, 0), 2)
                                        # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)

                                        self.path=r'%s'%self.fileName
                                        book = openpyxl.load_workbook(self.path)
                                        book_s=book.active
                                        nrows = book_s.max_row
                                        #print('表格總列數',nrows)
                                        book_s.cell(row=nrows+1,column=1).value = now_time
                                        book_s.cell(row=nrows+1,column=2).value = self.lineEdit.text()
                                        book_s.cell(row=nrows+1,column=3).value = self.lineEdit_2.text()
                                        book_s.cell(row=nrows+1,column=4).value = self.lineEdit_3.text()
                                        book_s.cell(row=nrows+1,column=5).value = self.lineEdit_4.text()
                                        book_s.cell(row=nrows+1,column=6).value = self.lineEdit_5.text()
                                        book_s.cell(row=nrows+1,column=7).value = self.lineEdit_7.text()
                                        book_s.cell(row=nrows+1,column=8).value = self.lineEdit_9.text()
                                        book_s.cell(row=nrows+1,column=9).value = self.lineEdit_6.text()
                                        book_s.cell(row=nrows+1,column=10).value = self.lineEdit_8.text()
                                        book_s.cell(row=nrows+1,column=11).value = now_H
                                        book_s.cell(row=nrows+1,column=12).value = now_S
                                        book_s.cell(row=nrows+1,column=13).value = now_V
                                        book_s.cell(row=nrows+1,column=14).value = "%d"%self.horizontalSlider_4.value()
                                        book_s.cell(row=nrows+1,column=15).value = "%d"%self.horizontalSlider.value()
                                        


                                        book_s.cell(row=nrows+1,column=16).value = now_S

                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                        book_s.cell(row=nrows+1,column=16).fill = fill

                                        book_s.cell(row=nrows+1,column=17).value = now_R
                                        book_s.cell(row=nrows+1,column=18).value = now_G
                                        book_s.cell(row=nrows+1,column=19).value = now_B
                                        book_s.cell(row=nrows+1, column=20).value = ('%s'%now_time)
                                        book_s.cell(row=nrows+1, column=20).hyperlink = ('%s'%self.pic_dir)
                                        book_s.cell(row=nrows+1, column=21).value = ('%s'%error_text)
                                        book.save('%s'%self.path)
                                        # cv2.putText(self.image, 'The picture and data have saved !',
                                        #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                        #                 cv2.FONT_HERSHEY_SIMPLEX,
                                        #                 1.0, (255, 0, 0), 2)
                                        # #now_time = time.strftime('%Y/%m/%d-%H:%M:%S',time.localtime(time.time()))
                                        # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)
                                        self.path_success = self.path
                                        self.path_success_count = 1
                                       # print ('Save completed')
                                        QMessageBox.information(None,"Tip","Save completed")

                                else:
                                        self.__file__=self.fileName
                                        self.pic_dir=os.path.dirname(self.__file__)+'/'+now_time+'.jpg'
                                        #print (pic_dir)
                                        self.path=r'%s'%self.fileName
                                        cv2.imencode('.jpg', now_image)[1].tofile(self.pic_dir)
                                        
                                        # cv2.putText(self.image, 'The picture and data have saved !',
                                        #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                        #                 cv2.FONT_HERSHEY_SIMPLEX,
                                        #                 1.0, (255, 0, 0), 2)
                                        # #now_time = time.strftime('%Y/%m/%d-%H:%M:%S',time.localtime(time.time()))
                                        # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)
                                        #self.timer_camera.stop()
                                        book = openpyxl.Workbook()
                                        sheet = book.active
                                        sheet.cell(row=1, column=1).value = "時間"
                                        sheet.cell(row=1, column=2).value = "產地"
                                        sheet.cell(row=1, column=3).value = "品種"
                                        sheet.cell(row=1, column=4).value = "栽培者"
                                        sheet.cell(row=1, column=5).value = "茶菁品質"
                                        sheet.cell(row=1, column=6).value = "重量"   
                                        sheet.cell(row=1, column=7).value = "目測萎凋程度(%)"
                                        sheet.cell(row=1, column=8).value = "實際水分含量(%)"
                                        sheet.cell(row=1, column=9).value = "備註1"
                                        sheet.cell(row=1, column=10).value = "備註2"
                                        sheet.cell(row=1, column=11).value = "色相(H)"
                                        sheet.cell(row=1, column=12).value = "飽和度(S)"
                                        sheet.cell(row=1, column=13).value = "亮度(V)"
                                        sheet.cell(row=1, column=14).value = "下限(S值)"
                                        sheet.cell(row=1, column=15).value = "上限(S值)"
                                        sheet.cell(row=1, column=16).value = "測量值(S值)"

                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                        sheet.cell(row=1,column=16).fill = fill

                                        sheet.cell(row=1, column=17).value = "R(紅)"
                                        sheet.cell(row=1, column=18).value = "G(綠)"
                                        sheet.cell(row=1, column=19).value = "B(藍)"
                                        sheet.cell(row=1, column=20).value = "圖片超連結"
                                        sheet.cell(row=1, column=21).value = "測量結果"
                                        
                                        sheet.cell(row=2, column=1).value = now_time
                                        sheet.cell(row=2, column=2).value = self.lineEdit.text()
                                        sheet.cell(row=2, column=3).value = self.lineEdit_2.text()
                                        sheet.cell(row=2, column=4).value = self.lineEdit_3.text()
                                        sheet.cell(row=2, column=5).value = self.lineEdit_4.text()
                                        sheet.cell(row=2, column=6).value = self.lineEdit_5.text()  
                                        sheet.cell(row=2, column=7).value = self.lineEdit_7.text()
                                        sheet.cell(row=2, column=8).value = self.lineEdit_9.text()
                                        sheet.cell(row=2, column=9).value = self.lineEdit_6.text()
                                        sheet.cell(row=2, column=10).value = self.lineEdit_8.text()
                                        sheet.cell(row=2, column=11).value = now_H
                                        sheet.cell(row=2, column=12).value = now_S
                                        sheet.cell(row=2, column=13).value = now_V
                                        sheet.cell(row=2, column=14).value = "%d"%self.horizontalSlider_4.value()
                                        sheet.cell(row=2, column=15).value = "%d"%self.horizontalSlider.value()
                                        sheet.cell(row=2, column=16).value = now_S

                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                        sheet.cell(row=2,column=16).fill = fill

                                        sheet.cell(row=2, column=17).value = now_R
                                        sheet.cell(row=2, column=18).value = now_G
                                        sheet.cell(row=2, column=19).value = now_B
                                        sheet.cell(row=2, column=20).value = ('%s'%now_time)
                                        sheet.cell(row=2, column=20).hyperlink = ('%s'%self.pic_dir)
                                        sheet.cell(row=2, column=21).value = ('%s'%error_text)
                                        book.save(self.fileName)
                                        self.path_success = self.fileName
                                        self.path_success_count = 1
                                # cv2.putText(self.image, 'The picture and data have saved !',
                                #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                #                 cv2.FONT_HERSHEY_SIMPLEX,
                                #                 1.0, (255, 0, 0), 2)
                                # #now_time = time.strftime('%Y/%m/%d-%H:%M:%S',time.localtime(time.time()))
                                # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)
                                        #print ('Save completed')
                                        QMessageBox.information(None,"Tip","Save completed")
                        
                        # show = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)  

                        # showImage = QtGui.QImage(show.data, show.shape[1], show.shape[0], QtGui.QImage.Format_RGB888)
                        # self.label_2.setPixmap(QtGui.QPixmap.fromImage(showImage))
                        # self.label_2.setScaledContents(True)

                
        else :
                QMessageBox.information(None,"Tip",r"Error: 上限值必須大於下限值")
                #print ('Error: 上限值必須大於下限值')


    def select(self):
        import imutils
        if self.horizontalSlider.value() >  self.horizontalSlider_4.value():
                self.select_check = 0
                self.check = -1
                cal_R=0
                cal_G=0
                cal_B=0
                cal_H=0
                cal_S=0
                cal_V=0 
                #cap = cv2.VideoCapture(0) 
                #flag,image = cap.read()
                #print (np.shape(self.image))
                #path = r'C:\Users\a0913\Desktop\2.jpg'   
                #image = cv2.imread(self.image)
                #image=cv2.flip(image,1)
                #print(np.shape(image))
                #show = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)
                self.now_image=self.image
                #print (np.shape(self.now_image))
                #print (np.shape(now_image))
                HSV = cv2.cvtColor(self.now_image, cv2.COLOR_BGR2HSV)
                (B, G, R) = cv2.split(self.now_image)
                (H, S, V) = cv2.split(HSV)
                #img=show
                self.now_image = imutils.resize(self.now_image, width=640)
                # print(np.shape(now_image))
                cv2.imshow("ROI", self.now_image)
                cv2.setMouseCallback("ROI", self.mouse_callback)
                print ("Press 'enter' to lock the ROI")
                key = cv2.waitKey(0) & 0xFF
                if key == 13 and self.check >0 :
                        print ("Press 'enter' to save the data")
                        self.select_check = 1
                        cv2.rectangle(img=self.img_last, pt1=(self.start_x, self.start_y), pt2=(self.x_last, self.y_last), color=(0, 0, 255), thickness=2)
                        cv2.imshow("ROI", self.img_last)
                        key = cv2.waitKey(0) & 0xFF
                        if key == 13 :
                                print ('save')
                                w = abs(self.x_last - self.start_x)
                                h = abs(self.y_last - self.start_y)
                                for Sx in range (self.start_x,self.start_x+w) :
                                        for Sy in range (self.start_y,self.start_y+h) :
                                                cal_R = cal_R + R[Sy][Sx]
                                                cal_G = cal_G + G[Sy][Sx]
                                                cal_B = cal_B + B[Sy][Sx]
                                                cal_H = cal_H + H[Sy][Sx]
                                                cal_S = cal_S + S[Sy][Sx]
                                                cal_V = cal_V + V[Sy][Sx]
                                if (w*h)!=0:  
                                        R_ROI=cal_R/(w*h)
                                        G_ROI=cal_G/(w*h)
                                        B_ROI=cal_B/(w*h)
                                        H_ROI=cal_H/(w*h)
                                        S_ROI=cal_S/(w*h)
                                        V_ROI=cal_V/(w*h)
                                # self.horizontalSlider_3.setSliderPosition(int(S_ROI))
                                        self.label_19.setText("%d"%(int(S_ROI)))
                                        Palette = QPalette()
                                        #color=QColor.fromHsv(h,s,v,a)
                                        Palette.setColor(QPalette.Window,QColor.fromHsv(116,(int(S_ROI)),116))#设置字体颜色
                                        self.label_19.setPalette(Palette)
                                                
                                        if self.horizontalSlider_4.value() == 0 :
                                                if int(S_ROI) > self.horizontalSlider.value() :
                                                        error = (int(S_ROI) - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                                        error = int(error)
                                                        error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                                        self.label_31.setText(error_text)
                                                else :
                                                        error_text = ('萎凋率0%')
                                                        self.label_31.setText(error_text)
                                        else :
                                                if int(S_ROI) > self.horizontalSlider.value() :
                                                        print ('too high')
                                                        error = (int(S_ROI) - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                                        error = int(error)
                                                        error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                                        self.label_31.setText(error_text)

                                                elif int(S_ROI) > self.horizontalSlider_4.value() and int(S_ROI) < self.horizontalSlider.value() :
                                                        print ('inside')
                                                        error_text = ('萎凋率0%')
                                                        self.label_31.setText(error_text)
                                                else :
                                                        print ('too low')
                                                        error = (int(S_ROI) - self.horizontalSlider_4.value())/self.horizontalSlider_4.value()*100
                                                        error = int(error)
                                                        error_text = ('萎凋率'+ '(' + '%s'%error + "%" + ')')
                                                        self.label_31.setText(error_text)

                                        now_time = time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
                                        now_time_no_hr = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                                        now_time_hr = time.strftime('%H-%M-%S',time.localtime(time.time()))
                                        #print(now_time)

                                        self.fileName,_ = QFileDialog.getSaveFileName(None,'save',now_time_no_hr,'Excel Files(*.xlsx)')
                                         # print (fileName)
                                        if self.fileName :
                                                if os.path.isfile(self.fileName):
                                                        self.__file__=self.fileName
                                                        self.pic_dir=os.path.dirname(self.__file__)+'/'+now_time+'.jpg'
                                                        #print (pic_dir)
                                                        cv2.imencode('.jpg', self.img_last)[1].tofile(self.pic_dir)
                                                        # cv2.putText(self.img_last, 'The picture and data have saved !',
                                                        #         (int(self.img_last.shape[1]/2-250), int(self.img_last.shape[0]/2)),
                                                        #         cv2.FONT_HERSHEY_SIMPLEX,
                                                        #         1.0, (255, 0, 0), 2)
                                                        # cv2.putText(self.img_last,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)

                                                        self.path=r'%s'%self.fileName
                                                        book = openpyxl.load_workbook(self.path)
                                                        book_s=book.active
                                                        nrows = book_s.max_row
                                                        #print('表格總列數',nrows)
                                                        book_s.cell(row=nrows+1,column=1).value = now_time
                                                        book_s.cell(row=nrows+1,column=2).value = self.lineEdit.text()
                                                        book_s.cell(row=nrows+1,column=3).value = self.lineEdit_2.text()
                                                        book_s.cell(row=nrows+1,column=4).value = self.lineEdit_3.text()
                                                        book_s.cell(row=nrows+1,column=5).value = self.lineEdit_4.text()
                                                        book_s.cell(row=nrows+1,column=6).value = self.lineEdit_5.text()
                                                        book_s.cell(row=nrows+1,column=7).value = self.lineEdit_7.text()
                                                        book_s.cell(row=nrows+1,column=8).value = self.lineEdit_9.text()
                                                        book_s.cell(row=nrows+1,column=9).value = self.lineEdit_6.text()
                                                        book_s.cell(row=nrows+1,column=10).value = self.lineEdit_8.text()
                                                        book_s.cell(row=nrows+1,column=11).value = int(H_ROI)
                                                        book_s.cell(row=nrows+1,column=12).value = int(S_ROI)
                                                        book_s.cell(row=nrows+1,column=13).value = int(V_ROI)
                                                        book_s.cell(row=nrows+1,column=14).value = "%d"%self.horizontalSlider_4.value()
                                                        book_s.cell(row=nrows+1,column=15).value = "%d"%self.horizontalSlider.value()
                                                        book_s.cell(row=nrows+1,column=16).value = int(S_ROI)

                                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                                        book_s.cell(row=nrows+1,column=16).fill = fill

                                                        book_s.cell(row=nrows+1,column=17).value = int(R_ROI)
                                                        book_s.cell(row=nrows+1,column=18).value = int(G_ROI)
                                                        book_s.cell(row=nrows+1,column=19).value = int(B_ROI)
                                                        book_s.cell(row=nrows+1, column=20).value = ('%s'%now_time)
                                                        book_s.cell(row=nrows+1, column=20).hyperlink = ('%s'%self.pic_dir)
                                                        book_s.cell(row=nrows+1, column=21).value = ('%s'%error_text)
                                                        book.save('%s'%self.path)
                                                        self.path_success = self.path
                                                        self.path_success_count = 1
                                                else:
                                                        self.__file__=self.fileName
                                                        self.pic_dir=os.path.dirname(self.__file__)+'/'+now_time+'.jpg'
                                                        #print (pic_dir)
                                                        cv2.imencode('.jpg', self.img_last)[1].tofile(self.pic_dir)
                                                        self.path=r'%s'%self.fileName
                                                        # cv2.putText(self.img_last, 'The picture and data have saved !',
                                                        #         (int(self.img_last.shape[1]/2-250), int(self.img_last.shape[0]/2)),
                                                        #         cv2.FONT_HERSHEY_SIMPLEX,
                                                        #         1.0, (255, 0, 0), 2)
                                                        # #now_time = time.strftime('%Y/%m/%d-%H:%M:%S',time.localtime(time.time()))
                                                        # cv2.putText(self.img_last,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)
                                                        #self.timer_camera.stop()
                                                        
                                                        book = openpyxl.Workbook()
                                                        sheet = book.active
                                                        sheet.cell(row=1, column=1).value = "時間"
                                                        sheet.cell(row=1, column=2).value = "產地"
                                                        sheet.cell(row=1, column=3).value = "品種"
                                                        sheet.cell(row=1, column=4).value = "栽培者"
                                                        sheet.cell(row=1, column=5).value = "茶菁品質"
                                                        sheet.cell(row=1, column=6).value = "重量"   
                                                        sheet.cell(row=1, column=7).value = "目測萎凋程度(%)"
                                                        sheet.cell(row=1, column=8).value = "實際水分含量(%)"
                                                        sheet.cell(row=1, column=9).value = "備註1"
                                                        sheet.cell(row=1, column=10).value = "備註2"
                                                        sheet.cell(row=1, column=11).value = "色相(H)"
                                                        sheet.cell(row=1, column=12).value = "飽和度(S)"
                                                        sheet.cell(row=1, column=13).value = "亮度(V)"
                                                        sheet.cell(row=1, column=14).value = "下限(S值)"
                                                        sheet.cell(row=1, column=15).value = "上限(S值)"
                                                        sheet.cell(row=1, column=16).value = "測量值(S值)"

                                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                                        sheet.cell(row=1,column=16).fill = fill


                                                        sheet.cell(row=1, column=17).value = "R(紅)"
                                                        sheet.cell(row=1, column=18).value = "G(綠)"
                                                        sheet.cell(row=1, column=19).value = "B(藍)"
                                                        sheet.cell(row=1, column=20).value = "圖片超連結"
                                                        sheet.cell(row=1, column=21).value = "測量結果"
                                                        
                                                        sheet.cell(row=2, column=1).value = now_time
                                                        sheet.cell(row=2, column=2).value = self.lineEdit.text()
                                                        sheet.cell(row=2, column=3).value = self.lineEdit_2.text()
                                                        sheet.cell(row=2, column=4).value = self.lineEdit_3.text()
                                                        sheet.cell(row=2, column=5).value = self.lineEdit_4.text()
                                                        sheet.cell(row=2, column=6).value = self.lineEdit_5.text()  
                                                        sheet.cell(row=2, column=7).value = self.lineEdit_7.text()
                                                        sheet.cell(row=2, column=8).value = self.lineEdit_9.text()
                                                        sheet.cell(row=2, column=9).value = self.lineEdit_6.text()
                                                        sheet.cell(row=2, column=10).value = self.lineEdit_8.text()
                                                        sheet.cell(row=2, column=11).value = int(H_ROI)
                                                        sheet.cell(row=2, column=12).value = int(S_ROI)
                                                        sheet.cell(row=2, column=13).value = int(V_ROI)
                                                        sheet.cell(row=2, column=14).value = "%d"%self.horizontalSlider_4.value()
                                                        sheet.cell(row=2, column=15).value = "%d"%self.horizontalSlider.value()
                                                        sheet.cell(row=2, column=16).value = int(S_ROI)

                                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                                        sheet.cell(row=2,column=16).fill = fill

                                                        sheet.cell(row=2, column=17).value = int(R_ROI)
                                                        sheet.cell(row=2, column=18).value = int(G_ROI)
                                                        sheet.cell(row=2, column=19).value = int(B_ROI)
                                                        sheet.cell(row=2, column=20).value = ('%s'%now_time)
                                                        sheet.cell(row=2, column=20).hyperlink = ('%s'%self.pic_dir)
                                                        sheet.cell(row=2, column=21).value = ('%s'%error_text)
                                                        book.save(self.fileName)
                                                        self.path_success = self.fileName
                                                        self.path_success_count = 1
                cv2.destroyAllWindows()                                                        
        else :
                QMessageBox.information(None,"Tip",r"Error: 上限值必須大於下限值")
                #print ('Error: 上限值必須大於下限值')
                            #cv2.imwrite(r'C:\Users\a0913\Desktop\1.jpg',now_image)

    def mouse_callback(self,event, x, y, flags, param):

        self.img_result = self.now_image.copy()

        if event == cv2.EVENT_LBUTTONDOWN and self.select_check == 0:
            self.check = 1

            self.mouse_is_pressing = True
            self.start_x, self.start_y = x,y

            cv2.circle(self.img_result, (x,y), 10, (255,0,0),-1)
           
            cv2.imshow("ROI", self.img_result)

        elif event == cv2.EVENT_MOUSEMOVE and self.select_check == 0:
            if self.mouse_is_pressing: 
                cv2.rectangle(self.img_result, (self.start_x, self.start_y), (x,y), (255,0,0), 2)
                cv2.imshow("ROI", self.img_result)

        elif event == cv2.EVENT_LBUTTONUP and self.select_check == 0:
            self.mouse_is_pressing = False 

            cv2.rectangle(img=self.img_result, pt1=(self.start_x, self.start_y), pt2=(x, y), color=(255, 0, 0), thickness=2)
            cv2.imshow("ROI", self.img_result)
           # print (self.start_x, self.start_y,x,y)        
            self.img_last=self.img_result
            self.x_last = x
            self.y_last = y

    def save(self):
        if self.horizontalSlider.value() >  self.horizontalSlider_4.value(): 
                #self.horizontalSlider_3.setSliderPosition(int(self.S_avg))
                now_H = self.H_avg
                now_H = int(now_H)
                now_S = self.S_avg
                now_S = int(now_S)
                now_V = self.V_avg
                now_V = int(now_V)
                now_R = self.R_avg
                now_R = int(now_R)
                now_G = self.G_avg
                now_G = int(now_G)
                now_B = self.B_avg
                now_B = int(now_B)

                self.label_19.setText(str(now_S))
                Palette = QPalette()
                #color=QColor.fromHsv(h,s,v,a)
                Palette.setColor(QPalette.Window,QColor.fromHsv(116,now_S,116))#设置字体颜色
                self.label_19.setPalette(Palette)

                if self.horizontalSlider_4.value() == 0 :
                        if now_S > self.horizontalSlider.value() :
                                error = (now_S - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)
                        else :
                                error_text = ('萎凋率0%')
                                self.label_31.setText(error_text)
                else :
                        if now_S > self.horizontalSlider.value() :
                                error = (now_S - self.horizontalSlider.value())/self.horizontalSlider.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + "+" + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)

                        elif now_S >= self.horizontalSlider_4.value() and now_S <= self.horizontalSlider.value() :
                                error_text = ('萎凋率0%')
                                self.label_31.setText(error_text)
                        else :
                                error = (now_S - self.horizontalSlider_4.value())/self.horizontalSlider_4.value()*100
                                error = int(error)
                                error_text = ('萎凋率'+ '(' + '%s'%error + "%" + ')')
                                self.label_31.setText(error_text)

                if self.timer_camera.isActive() != False:
                        now_time = time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
                        now_time_no_hr = time.strftime('%Y-%m-%d',time.localtime(time.time()))
                        now_time_hr = time.strftime('%H-%M-%S',time.localtime(time.time()))
                        #print(now_time)
                        # fileName,_ = QFileDialog.getSaveFileName(None,'save',now_time_no_hr,'Excel Files(*.xlsx)')
                        # print (fileName)
                        if self.path_success_count == 1 :
                                if os.path.isfile(self.path_success):
                                        self.__file__=self.path_success
                                        self.pic_dir=os.path.dirname(self.__file__)+'/'+now_time+'.jpg'
                                        #print (pic_dir)
                                        cv2.imencode('.jpg', self.image)[1].tofile(self.pic_dir)
                                        # cv2.putText(self.image, 'The picture and data have saved !',
                                        #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                        #                 cv2.FONT_HERSHEY_SIMPLEX,
                                        #                 1.0, (255, 0, 0), 2)
                                        # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)

                                        self.path=r'%s'%self.path_success
                                        book = openpyxl.load_workbook(self.path)
                                        book_s=book.active
                                        nrows = book_s.max_row
                                        #print('表格總列數',nrows)
                                        book_s.cell(row=nrows+1,column=1).value = now_time
                                        book_s.cell(row=nrows+1,column=2).value = self.lineEdit.text()
                                        book_s.cell(row=nrows+1,column=3).value = self.lineEdit_2.text()
                                        book_s.cell(row=nrows+1,column=4).value = self.lineEdit_3.text()
                                        book_s.cell(row=nrows+1,column=5).value = self.lineEdit_4.text()
                                        book_s.cell(row=nrows+1,column=6).value = self.lineEdit_5.text()
                                        book_s.cell(row=nrows+1,column=7).value = self.lineEdit_7.text()
                                        book_s.cell(row=nrows+1,column=8).value = self.lineEdit_9.text()
                                        book_s.cell(row=nrows+1,column=9).value = self.lineEdit_6.text()
                                        book_s.cell(row=nrows+1,column=10).value = self.lineEdit_8.text()
                                        book_s.cell(row=nrows+1,column=11).value = now_H
                                        book_s.cell(row=nrows+1,column=12).value = now_S
                                        book_s.cell(row=nrows+1,column=13).value = now_V
                                        book_s.cell(row=nrows+1,column=14).value = "%d"%self.horizontalSlider_4.value()
                                        book_s.cell(row=nrows+1,column=15).value = "%d"%self.horizontalSlider.value()

                                        fill = PatternFill("solid", fgColor="AAFFEE")
                                        book_s.cell(row=nrows+1,column=16).fill = fill


                                        book_s.cell(row=nrows+1,column=16).value = now_S
                                        book_s.cell(row=nrows+1,column=17).value = now_R
                                        book_s.cell(row=nrows+1,column=18).value = now_G
                                        book_s.cell(row=nrows+1,column=19).value = now_B
                                        book_s.cell(row=nrows+1, column=20).value = ('%s'%now_time)
                                        book_s.cell(row=nrows+1, column=20).hyperlink = ('%s'%self.pic_dir)
                                        book_s.cell(row=nrows+1, column=21).value = ('%s'%error_text)
                                        book.save('%s'%self.path)
                                        # cv2.putText(self.image, 'The picture and data have saved !',
                                        #                 (int(self.image.shape[1]/2-250), int(self.image.shape[0]/2)),
                                        #                 cv2.FONT_HERSHEY_SIMPLEX,
                                        #                 1.0, (255, 0, 0), 2)
                                        # #now_time = time.strftime('%Y/%m/%d-%H:%M:%S',time.localtime(time.time()))
                                        # cv2.putText(self.image,now_time,(20,160),cv2.FONT_HERSHEY_SIMPLEX,1,(0,0,0),2)
                                        QMessageBox.information(None,"Tip","Save completed")
                                        #print ('Save completed')
                        else:
                                #print ('Please to load or create a excel file')
                                QMessageBox.information(None,"Tip","Please to load or create a excel file")
                        # show = cv2.cvtColor(self.image, cv2.COLOR_BGR2RGB)  

                        # showImage = QtGui.QImage(show.data, show.shape[1], show.shape[0], QtGui.QImage.Format_RGB888)
                        # self.label_2.setPixmap(QtGui.QPixmap.fromImage(showImage))
                        # self.label_2.setScaledContents(True)

        else :
               # print ('Error: 上限值必須大於下限值')
                QMessageBox.information(None,"Tip",r"Error: 上限值必須大於下限值")
    def closeEvent(self):
        #print ('event')
        led = digitalio.DigitalInOut(board.C0)
        led.direction = digitalio.Direction.OUTPUT
        led.value = False
        

if __name__ == '__main__': 

    app_progress = QApplication(sys.argv)
    window = Actions()
    QApplication.closeAllWindows()

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow(MainWindow)
    #ui.setupUi(MainWindow)
    MainWindow.setWindowTitle('Tea withering system')
    MainWindow.show()

    def update_label(): 
         current_time = time.strftime("%Y-%m-%d %H:%M:%S") 
         ui.label_4.setText(current_time)

    timer = QtCore.QTimer() 
    timer.timeout.connect(update_label) 
    timer.start(1000) # every 1,000 milliseconds

    sys.exit(app.exec_()) 
